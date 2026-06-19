"""Workbook zero-error gate (VNDP golden rule).

Scans the repaired workbook for any spreadsheet calculation error
(`#REF!`, `#VALUE!`, `#DIV/0!`, `#NAME?`, `#N/A`, `#NULL!`, `#NUM!`) in cached cell
values and exits non-zero if any are found. This is wired into CI (see .github/workflows/ci.yml)
and is the systemic guard against the `#REF!`-class defect that the platform exists to cure.

Usage (always `python`, never `python3`):
    python workbook/verify.py
    python workbook/verify.py --file "workbook/22 Cr Nutracare Project - REPAIRED.xlsx"

Primary mode uses openpyxl on cached values (fast, no LibreOffice/recalc dependency).
The repaired workbook ships verified to 0 errors; this asserts it stays that way.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ERROR_LITERALS = (
    "#REF!",
    "#VALUE!",
    "#DIV/0!",
    "#NAME?",
    "#N/A",
    "#NULL!",
    "#NUM!",
)

DEFAULT_WORKBOOK = Path(__file__).with_name("22 Cr Nutracare Project - REPAIRED.xlsx")


def scan(path: Path) -> dict[str, list[str]]:
    """Return {sheet_name: [cell_refs...]} for every cell holding an error literal."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - environment guard
        sys.stderr.write("openpyxl is required: `uv pip install openpyxl`\n")
        raise SystemExit(2) from None

    if not path.exists():
        sys.stderr.write(f"Workbook not found: {path}\n")
        raise SystemExit(2)

    # data_only=True reads cached computed values; that is where error literals surface.
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    findings: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        bad: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in ERROR_LITERALS:
                    bad.append(f"{cell.coordinate}={value}")
        if bad:
            findings[ws.title] = bad
    wb.close()
    return findings


def main() -> int:
    parser = argparse.ArgumentParser(description="VNDP workbook zero-error gate")
    parser.add_argument("--file", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    findings = scan(args.file)
    total = sum(len(v) for v in findings.values())

    if total == 0:
        print(f"OK: 0 calculation errors in {args.file.name}")
        return 0

    print(f"FAIL: {total} calculation error(s) in {args.file.name}", file=sys.stderr)
    for sheet, cells in sorted(findings.items()):
        preview = ", ".join(cells[:10]) + (" …" if len(cells) > 10 else "")
        print(f"  [{sheet}] {len(cells)}: {preview}", file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
